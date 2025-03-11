import pandas as pd
from openpyxl import load_workbook
import openpyxl.utils

class ExcelReader:
    def __init__(self, file_name, sheet_name, cell_address):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.cell_address = cell_address

    def cell_to_position(self, cell_address):
        """
        셀 주소(예: 'A3')를 (행, 열) 숫자 인덱스로 변환합니다.
        :param cell_address: 셀 주소 (예: 'A3')
        :return: (row, column)
        """
        try:
            column = openpyxl.utils.cell.column_index_from_string(cell_address[:1])
            row = int(cell_address[1:])
            return row, column
        except Exception as e:
            raise ValueError(f"셀 주소를 변환하는 중 오류 발생: {e}")

    def read_to_sheet(self):
        """
        엑셀 파일에서 데이터를 읽어 DataFrame으로 반환합니다.
        :return: (성공 여부, DataFrame 또는 오류 메시지)
        """
        try:
            # 셀 주소를 시작 행/열로 변환
            start_row, start_col = self.cell_to_position(self.cell_address)

            # 엑셀 파일 로드
            wb = load_workbook(self.file_name, data_only=True)
            if self.sheet_name not in wb.sheetnames:
                raise ValueError(f"시트 '{self.sheet_name}'이(가) 존재하지 않습니다.")

            sheet = wb[self.sheet_name]

            # 데이터 읽기
            data = []
            for row in sheet.iter_rows(min_row=start_row, min_col=start_col, values_only=True):
                data.append(row)

            if not data:
                raise ValueError("지정된 셀 범위에 데이터가 없습니다.")

            # DataFrame 생성
            df = pd.DataFrame(data[1:], columns=data[0])
            return True, df

        except FileNotFoundError:
            return False, f"파일 '{self.file_name}'이(가) 존재하지 않습니다."
        except ValueError as ve:
            return False, f"ValueError: {ve}"
        except Exception as e:
            return False, f"알 수 없는 오류 발생: {e}"

if __name__ == "__main__":
    file_name = "your_file.xlsx"
    sheet_name = "Sheet1"
    cell_address = "A2"

    reader = ExcelReader(file_name, sheet_name, cell_address)
    success, result = reader.read_to_sheet()

    if success:
        print(result.head())
    else:
        print(result)
